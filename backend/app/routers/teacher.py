import os
import io
import chardet
from datetime import datetime
from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from pydantic import BaseModel
from app.database import db
from app.auth_utils import create_token, verify_teacher_token
from pypinyin import lazy_pinyin, Style
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

TEACHER_PASSWORD = os.environ.get("TEACHER_PASSWORD", "admin123")


def _require_teacher(authorization: Optional[str]):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="请先登录")
    token = authorization.removeprefix("Bearer ").strip()
    try:
        verify_teacher_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))


def _name_to_pinyin(name: str):
    full = "".join(lazy_pinyin(name, style=Style.NORMAL))
    abbr = "".join(lazy_pinyin(name, style=Style.FIRST_LETTER))
    return full.lower(), abbr.lower()


class LoginRequest(BaseModel):
    password: str


@router.post("/api/teacher/login")
def teacher_login(req: LoginRequest):
    if req.password != TEACHER_PASSWORD:
        raise HTTPException(status_code=401, detail="密码错误")
    token = create_token({"role": "teacher"}, expires_hours=8)
    return {"token": token}


@router.post("/api/teacher/students")
async def upload_students(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """上传学生名单 CSV（UTF-8 或 GBK 均支持）"""
    _require_teacher(authorization)

    raw = await file.read()
    encoding = chardet.detect(raw)["encoding"] or "utf-8"
    text = raw.decode(encoding, errors="replace")

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        raise HTTPException(status_code=422, detail="文件为空")

    # 跳过表头
    data_lines = lines[1:] if lines[0].startswith("姓名") or lines[0].startswith("name") else lines

    records = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 2:
            continue
        name = parts[0]
        student_id = parts[1]
        class_name = parts[2] if len(parts) > 2 else ""
        if not name or not student_id:
            continue
        pinyin, abbr = _name_to_pinyin(name)
        records.append((name, student_id, class_name, pinyin, abbr))

    if not records:
        raise HTTPException(status_code=422, detail="CSV 中没有有效数据行")

    with db() as conn:
        conn.executemany("""
            INSERT INTO students (name, student_id, class_name, pinyin, pinyin_abbr)
            VALUES (?,?,?,?,?)
            ON CONFLICT(student_id) DO UPDATE SET
                name=excluded.name,
                class_name=excluded.class_name,
                pinyin=excluded.pinyin,
                pinyin_abbr=excluded.pinyin_abbr
        """, records)

    return {"count": len(records)}


class AddStudentRequest(BaseModel):
    name: str
    student_id: str
    class_name: str = ""


@router.post("/api/teacher/students/add")
def add_student(req: AddStudentRequest, authorization: Optional[str] = Header(None)):
    """添加单个学生"""
    _require_teacher(authorization)
    req.name = req.name.strip()
    req.student_id = req.student_id.strip()
    if not req.name or not req.student_id:
        raise HTTPException(status_code=422, detail="姓名和学号不能为空")
    pinyin, abbr = _name_to_pinyin(req.name)
    with db() as conn:
        existing = conn.execute(
            "SELECT name FROM students WHERE student_id=?", (req.student_id,)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail=f"学号 {req.student_id} 已存在（{existing['name']}）")
        conn.execute(
            "INSERT INTO students (name, student_id, class_name, pinyin, pinyin_abbr) VALUES (?,?,?,?,?)",
            (req.name, req.student_id, req.class_name.strip(), pinyin, abbr)
        )
    return {"ok": True}


@router.delete("/api/teacher/students/{student_id}")
def delete_student(student_id: str, authorization: Optional[str] = Header(None)):
    """删除单个学生（同时删除其成绩）"""
    _require_teacher(authorization)
    with db() as conn:
        student = conn.execute(
            "SELECT name FROM students WHERE student_id=?", (student_id,)
        ).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="学号不存在")
        conn.execute("DELETE FROM scores WHERE student_id=?", (student_id,))
        conn.execute("DELETE FROM students WHERE student_id=?", (student_id,))
    return {"ok": True, "deleted": student["name"]}


@router.delete("/api/teacher/students")
def clear_all_students(authorization: Optional[str] = Header(None)):
    """清空全部学生名单（同时清空所有成绩）"""
    _require_teacher(authorization)
    with db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        conn.execute("DELETE FROM scores")
        conn.execute("DELETE FROM students")
    return {"ok": True, "deleted_count": count}


@router.post("/api/teacher/reset")
def new_semester_reset(authorization: Optional[str] = Header(None)):
    """新学期重置：清空所有学生名单 + 所有成绩"""
    _require_teacher(authorization)
    with db() as conn:
        students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        scores   = conn.execute("SELECT COUNT(*) FROM scores").fetchone()[0]
        conn.execute("DELETE FROM scores")
        conn.execute("DELETE FROM students")
    return {"ok": True, "deleted_students": students, "deleted_scores": scores}


@router.get("/api/teacher/students/list")
def list_students(authorization: Optional[str] = Header(None)):
    """获取全部学生名单"""
    _require_teacher(authorization)
    with db() as conn:
        rows = conn.execute(
            "SELECT name, student_id, class_name FROM students ORDER BY class_name, name"
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/api/teacher/exams")
def list_exams(authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        exams = conn.execute("SELECT id, title, is_active FROM exams ORDER BY id").fetchall()
        total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        result = []
        for e in exams:
            submitted = conn.execute(
                "SELECT COUNT(*) FROM scores WHERE exam_id=?", (e["id"],)
            ).fetchone()[0]
            avg = conn.execute(
                "SELECT AVG(score) FROM scores WHERE exam_id=?", (e["id"],)
            ).fetchone()[0]
            result.append({
                "id": e["id"], "title": e["title"], "is_active": e["is_active"],
                "submitted": submitted, "total_students": total_students,
                "avg_score": round(avg, 1) if avg else None,
            })
    return result


class ExamCreate(BaseModel):
    id: str
    title: str


@router.post("/api/teacher/exams")
def create_exam(req: ExamCreate, authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO exams (id, title) VALUES (?,?)",
            (req.id, req.title)
        )
    return {"ok": True}


class ExamUpdate(BaseModel):
    is_active: int


@router.put("/api/teacher/exams/{exam_id}")
def update_exam(exam_id: str, req: ExamUpdate, authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        conn.execute("UPDATE exams SET is_active=? WHERE id=?", (req.is_active, exam_id))
    return {"ok": True}


@router.get("/api/teacher/scores")
def get_scores(exam_id: str = Query(...), authorization: Optional[str] = Header(None)):
    """获取某次考试的所有学生成绩（含未提交）"""
    _require_teacher(authorization)
    with db() as conn:
        exam = conn.execute("SELECT title FROM exams WHERE id=?", (exam_id,)).fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        students = conn.execute(
            "SELECT name, student_id, class_name FROM students ORDER BY student_id"
        ).fetchall()
        scores_map = {
            r["student_id"]: dict(r)
            for r in conn.execute(
                "SELECT student_id, score, total, submitted_at FROM scores WHERE exam_id=?",
                (exam_id,)
            ).fetchall()
        }

    result = []
    for s in students:
        sc = scores_map.get(s["student_id"])
        result.append({
            "name": s["name"],
            "student_id": s["student_id"],
            "class_name": s["class_name"],
            "score": sc["score"] if sc else None,
            "total": sc["total"] if sc else None,
            "submitted_at": sc["submitted_at"] if sc else None,
        })

    return {"exam_title": exam["title"], "rows": result}


@router.get("/api/teacher/scores/export")
def export_scores(exam_id: str = Query(...), authorization: Optional[str] = Header(None)):
    """导出成绩 Excel"""
    _require_teacher(authorization)

    with db() as conn:
        exam = conn.execute("SELECT title FROM exams WHERE id=?", (exam_id,)).fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        students = conn.execute(
            "SELECT name, student_id, class_name FROM students ORDER BY student_id"
        ).fetchall()
        scores_map = {
            r["student_id"]: dict(r)
            for r in conn.execute(
                "SELECT student_id, score, total, submitted_at FROM scores WHERE exam_id=?",
                (exam_id,)
            ).fetchall()
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = exam["title"][:31]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["姓名", "学号", "班级", "得分", "满分", "提交时间"]
    col_widths = [12, 14, 20, 8, 8, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for row_idx, s in enumerate(students, 2):
        sc = scores_map.get(s["student_id"])
        ws.cell(row=row_idx, column=1, value=s["name"])
        ws.cell(row=row_idx, column=2, value=s["student_id"])
        ws.cell(row=row_idx, column=3, value=s["class_name"])
        ws.cell(row=row_idx, column=4, value=sc["score"] if sc else "")
        ws.cell(row=row_idx, column=5, value=sc["total"] if sc else "")
        ws.cell(row=row_idx, column=6, value=sc["submitted_at"] if sc else "")
        if not sc:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).font = Font(color="999999")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"成绩单_{exam['title']}_{date_str}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
